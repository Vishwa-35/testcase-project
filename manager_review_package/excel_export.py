"""
Excel Export Module

EXPORT STRUCTURE:
- Sheet 1: Dashboard (Project Overview + Status Summary)
- Sheet 2: History (Historical execution records)
- Sheet 3+: Test Case Sheets (one per sheet_name)
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from functools import reduce
from operator import or_

from django.db.models import Q, Case, When, IntegerField, Value
from django.db.models.functions import Cast

from .models import TestCase, SheetMeta, TestExecution, ProjectOverview, TestExecutionSnapshot
from testmanager.constants import EXCEL_COLUMN_WIDTH
from testmanager.version_service import get_base_test_case_id
from testmanager.services import build_search_filters, get_active_instance
from testmanager.utils import get_requirement_id


# =====================================================
# CONSTANTS & STYLES
# =====================================================

HEADERS = [
    "SL.NO", "SW Part Number", "Feature", "Requirement ID",
    "Requirement Description", "Test Case ID", "Test Case Summary",
    "Pre Conditions", "Inputs", "Periodic Time", "Test Steps",
    "Expected Result", "Status", "Reports", "Comments"
]

FIXED_WIDTH = EXCEL_COLUMN_WIDTH

CENTER_WRAP = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True
)

HEADER_FILL = PatternFill("solid", fgColor="0052CC")  # Blue
HEADER_FONT = Font(color="FFFFFF", bold=True)

FEATURE_FILL = PatternFill("solid", fgColor="FFF3B0")  # Light Yellow

STATUS_PASS_FILL = PatternFill("solid", fgColor="C6EFCE")  # Green
STATUS_FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")  # Red
STATUS_NOT_RELEVANT_FILL = PatternFill("solid", fgColor="FFEB9C")  # Yellow
STATUS_NOT_EXECUTED_FILL = PatternFill("solid", fgColor="D3D3D3")  # Gray

SUMMARY_PASS_FILL = PatternFill("solid", fgColor="C6EFCE")  # Green
SUMMARY_FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")  # Red
SUMMARY_NOT_RELEVANT_FILL = PatternFill("solid", fgColor="FFEB9C")  # Yellow
SUMMARY_TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")  # Light Blue
SUMMARY_EXECUTED_FILL = PatternFill("solid", fgColor="D9E1F2")  # Light Blue

TITLE_FONT = Font(bold=True, size=16)
TITLE_ALIGN = Alignment(horizontal="center", vertical="center")
SUMMARY_LABEL_FONT = Font(bold=True)


# =====================================================
# HELPER: GET MOST RECENTLY CREATED VERSION
# =====================================================

def _get_most_recently_created_version(sw_part_numbers):
    """
    For each SW Part Number, get the active version (is_active=True).
    Returns dict: {sw_part_number: version}
    """
    from .models import SWVersionMapping
    
    # Get active instance - only work with active instance data
    active_instance = get_active_instance()
    
    latest_versions = {}
    
    for sw_num in sw_part_numbers:
        if not sw_num:
            continue
        
        # Get active version from SWVersionMapping
        mapping = SWVersionMapping.objects.filter(
            instance=active_instance,
            sw_part_number=sw_num,
            is_active=True
        ).first()
        
        if mapping:
            latest_versions[sw_num] = mapping.version
    
    return latest_versions


# =====================================================
# HELPER: WRITE DASHBOARD SHEET
# =====================================================

def _write_dashboard_sheet(wb, sheet_names, version_selections, active_instance=None):
    """
    Create Sheet 1: Dashboard
    - Title merged rows 1-3, columns A-O
    - Project Overview section starting row 7, columns E-F
    - Status Summary starting row 7, columns M-P
    """
    from functools import reduce
    from operator import or_
    from django.db.models import Q
    
    if active_instance is None:
        active_instance = get_active_instance()
    
    ws = wb.create_sheet(title="Dashboard", index=0)
    
    # Title: Merge rows 1-3, columns A-O
    ws.merge_cells('A1:O3')
    title_cell = ws['A1']
    title_cell.value = "Dashboard"
    title_cell.font = TITLE_FONT
    title_cell.alignment = TITLE_ALIGN
    
    # Project Overview Section
    # Header row 7, column E
    ws['E7'] = "Project Overview"
    ws['E7'].font = Font(bold=True)
    
    # Get ProjectOverview data from database (use first available entry or create defaults)
    project_overview_obj = ProjectOverview.objects.filter(instance=active_instance).exclude(key__in=['last_export_timestamp']).first()
    
    # Extract Project Overview values
    project_code = project_overview_obj.project_code if project_overview_obj else ""
    vcu_platform = project_overview_obj.vcu_platform if project_overview_obj else ""
    hw_part_number = project_overview_obj.hardware_part_number if project_overview_obj else ""
    project_stage = project_overview_obj.project_stage if project_overview_obj else ""
    developer = project_overview_obj.developer if project_overview_obj else ""
    test_engineer = project_overview_obj.test_engineer if project_overview_obj else ""
    bootloader_sw_version = project_overview_obj.bootloader_sw_version if project_overview_obj else ""
    checksum_value = project_overview_obj.checksum_value if project_overview_obj else ""
    dbc_test_it = project_overview_obj.dbc_test_it if project_overview_obj else ""
    
    # Software Part Number: ALL SWs (comma-separated)
    # Collect unique SW part numbers
    sw_part_numbers_set = set()
    for sw_num, version in version_selections.items():
        sw_part_numbers_set.add(sw_num)
    sw_part_numbers_list = sorted(list(sw_part_numbers_set))
    software_part_number = ", ".join(sw_part_numbers_list) if sw_part_numbers_list else ""
    
    # Application SW Version: versions selected per SW (comma-separated)
    # Handle both single version strings and lists of versions
    app_sw_versions_list = []
    for sw, version in sorted(version_selections.items()):
        if isinstance(version, list):
            # Multiple versions for same SW
            versions_str = ", ".join(version)
            app_sw_versions_list.append(f"{sw}: {versions_str}")
        else:
            app_sw_versions_list.append(f"{sw}: {version}")
    application_sw_version = ", ".join(app_sw_versions_list) if app_sw_versions_list else ""
    
    # Project Overview fields (EXACT order as specified)
    overview_data = [
        ("Project Code", project_code),
        ("VCU Platform", vcu_platform),
        ("Hardware Part Number", hw_part_number),
        ("Software Part Number", software_part_number),
        ("Project Stage", project_stage),
        ("Developer", developer),
        ("Test Engineer", test_engineer),
        ("Application SW Version", application_sw_version),
        ("Bootloader SW Version", bootloader_sw_version),
        ("Checksum Value", checksum_value),
        ("DBC test_it", dbc_test_it),
    ]
    
    # Write Project Overview (starting row 8)
    row = 8
    for label, value in overview_data:
        ws.cell(row=row, column=5).value = label  # Column E
        ws.cell(row=row, column=6).value = value  # Column F
        row += 1
    
    # Status Summary Section
    # Header row 7, column M
    ws['M7'] = "Status Summary"
    ws['M7'].font = Font(bold=True)
    
    # Calculate status summary PER SHEET (Sheet 3+)
    # Get all test cases for all sheets (active instance only)
    all_test_cases = TestCase.objects.filter(instance=active_instance, sheet_name__in=sheet_names)
    
    # Filter test cases to only selected versions per SW Part Number
    # Handle both single version strings and lists of versions
    version_filters_list = []
    for sw_num, version in version_selections.items():
        if isinstance(version, list):
            # Multiple versions for same SW
            for v in version:
                version_filters_list.append(Q(sw_part_number=sw_num) & Q(app_sw_version=v))
        else:
            version_filters_list.append(Q(sw_part_number=sw_num) & Q(app_sw_version=version))
    
    if version_filters_list:
        version_filter = reduce(or_, version_filters_list)
        all_test_cases = all_test_cases.filter(version_filter)
    
    # Get executions for selected versions only (active instance only)
    test_case_ids = list(all_test_cases.values_list('id', flat=True))
    executions = TestExecution.objects.filter(instance=active_instance, test_case_id__in=test_case_ids)
    
    # Filter executions to selected versions
    exec_version_filters_list = []
    for sw_num, version in version_selections.items():
        if isinstance(version, list):
            # Multiple versions for same SW
            for v in version:
                exec_version_filters_list.append(Q(sw_part_number=sw_num) & Q(app_sw_version=v))
        else:
            exec_version_filters_list.append(Q(sw_part_number=sw_num) & Q(app_sw_version=version))
    
    if exec_version_filters_list:
        exec_version_filter = reduce(or_, exec_version_filters_list)
        executions = executions.filter(exec_version_filter)
    
    # Count by status
    pass_count = executions.filter(status__iexact="pass").count()
    fail_count = executions.filter(status__iexact="fail").count()
    not_relevant_count = executions.filter(
        Q(status__iexact="not relevant") | Q(status__iexact="not_relevant")
    ).count()
    total_test_cases = all_test_cases.count()
    not_executed_count = total_test_cases - pass_count - fail_count - not_relevant_count
    
    # Write Status Summary (starting row 8)
    ws.cell(row=8, column=13).value = "PASS"  # Column M
    ws.cell(row=8, column=14).value = pass_count  # Column N
    ws.cell(row=9, column=13).value = "FAIL"  # Column M
    ws.cell(row=9, column=14).value = fail_count  # Column N
    ws.cell(row=10, column=13).value = "NOT RELEVANT"  # Column M
    ws.cell(row=10, column=14).value = not_relevant_count  # Column N
    ws.cell(row=11, column=13).value = "NOT EXECUTED"  # Column M
    ws.cell(row=11, column=14).value = not_executed_count  # Column N
    
    # Set column widths
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 15


# =====================================================
# HELPER: WRITE HISTORY SHEET
# =====================================================

def _write_history_sheet(wb):
    """
    Create Sheet 2: History
    - Title merged rows 1-3, columns A-O
    - History table starting row 7, columns E-I
    """
    ws = wb.create_sheet(title="History", index=1)
    
    # Title: Merge rows 1-3, columns A-O
    ws.merge_cells('A1:O3')
    title_cell = ws['A1']
    title_cell.value = "History"
    title_cell.font = TITLE_FONT
    title_cell.alignment = TITLE_ALIGN
    
    # History Table Headers (row 7)
    headers = ["Test Case ID", "SW Part Number", "Version", "Status", "Executed At"]
    for col_idx, header in enumerate(headers, start=5):  # Start at column E
        cell = ws.cell(row=7, column=col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
    
    # Get historical execution records from TestExecutionSnapshot
    snapshots = TestExecutionSnapshot.objects.all().order_by('-exported_at')[:100]  # Last 100 snapshots
    
    row = 8
    for snapshot in snapshots:
        for exec_data in snapshot.execution_data[:50]:  # Limit per snapshot
            # Use base_test_case_id from snapshot data (version suffix never shown in exports)
            test_case_id = exec_data.get('test_case_id', '')
            # Try to extract base_test_case_id if version suffix exists
            base_test_case_id = exec_data.get('base_test_case_id', '')
            if not base_test_case_id and test_case_id:
                # Fallback: try to remove version suffix
                version = exec_data.get('execution_app_sw_version', '') or exec_data.get('app_sw_version', '')
                if version and test_case_id.endswith(f"_{version}"):
                    base_test_case_id = test_case_id[:-len(f"_{version}")]
                else:
                    base_test_case_id = test_case_id
            
            sw_part_number = exec_data.get('sw_part_number', '')
            version = exec_data.get('execution_app_sw_version', '') or exec_data.get('app_sw_version', '')
            status = exec_data.get('execution_status', '') or exec_data.get('status', '')
            executed_at = exec_data.get('executed_at', '')
            
            ws.cell(row=row, column=5).value = base_test_case_id or test_case_id  # Column E - always base_test_case_id
            ws.cell(row=row, column=6).value = sw_part_number  # Column F
            ws.cell(row=row, column=7).value = version  # Column G
            ws.cell(row=row, column=8).value = status  # Column H
            ws.cell(row=row, column=9).value = executed_at  # Column I
            row += 1
    
    # Set column widths
    for col in range(5, 10):  # Columns E-I
        ws.column_dimensions[get_column_letter(col)].width = 20


# =====================================================
# HELPER: WRITE TEST CASE SHEET DATA
# =====================================================

def _write_test_case_sheet(ws, qs, latest_versions, active_instance=None):
    """
    Write test case data to worksheet with new layout:
    - Rows 1-6: Empty
    - Rows 8-9: Summary values (columns A-E)
    - Charts in columns F-G (pie) and H-I (bar)
    - Row 10: Header (Blue background, white text, filters enabled)
    - Row 11+: Data
    - Feature column: Light Yellow
    - Status column: Color-coded (PASS=green, FAIL=red, NOT RELEVANT=yellow, NOT EXECUTED=gray)
    """
    if active_instance is None:
        active_instance = get_active_instance()
    
    total_test_cases = qs.count()
    pass_count = 0
    fail_count = 0
    not_relevant_count = 0
    executed_count = 0
    
    if qs.exists():
        test_case_ids = list(qs.values_list('id', flat=True))
        executions = TestExecution.objects.filter(instance=active_instance, test_case_id__in=test_case_ids).select_related('test_case', 'user')
        
        exec_version_filters_list = [
            Q(sw_part_number=sw_num) & Q(app_sw_version=version)
            for sw_num, version in latest_versions.items()
        ]
        if exec_version_filters_list:
            exec_version_filter = reduce(or_, exec_version_filters_list)
            executions = executions.filter(exec_version_filter)
        
        pass_count = executions.filter(status__iexact="pass").count()
        fail_count = executions.filter(status__iexact="fail").count()
        not_relevant_count = executions.filter(
            Q(status__iexact="not relevant") | Q(status__iexact="not_relevant")
        ).count()
        executed_count = pass_count + fail_count + not_relevant_count
    
    summary_row_labels = 8
    summary_row_values = 9
    
    ws.cell(row=summary_row_labels, column=1).value = "Total Testcases"
    ws.cell(row=summary_row_values, column=1).value = total_test_cases
    ws.cell(row=summary_row_labels, column=1).fill = SUMMARY_TOTAL_FILL
    ws.cell(row=summary_row_values, column=1).fill = SUMMARY_TOTAL_FILL
    ws.cell(row=summary_row_labels, column=1).font = SUMMARY_LABEL_FONT
    ws.cell(row=summary_row_labels, column=1).alignment = CENTER_WRAP
    ws.cell(row=summary_row_values, column=1).alignment = CENTER_WRAP
    
    ws.cell(row=summary_row_labels, column=2).value = "Passed"
    ws.cell(row=summary_row_values, column=2).value = pass_count
    ws.cell(row=summary_row_labels, column=2).fill = SUMMARY_PASS_FILL
    ws.cell(row=summary_row_values, column=2).fill = SUMMARY_PASS_FILL
    ws.cell(row=summary_row_labels, column=2).font = SUMMARY_LABEL_FONT
    ws.cell(row=summary_row_labels, column=2).alignment = CENTER_WRAP
    ws.cell(row=summary_row_values, column=2).alignment = CENTER_WRAP
    
    ws.cell(row=summary_row_labels, column=3).value = "Failed"
    ws.cell(row=summary_row_values, column=3).value = fail_count
    ws.cell(row=summary_row_labels, column=3).fill = SUMMARY_FAIL_FILL
    ws.cell(row=summary_row_values, column=3).fill = SUMMARY_FAIL_FILL
    ws.cell(row=summary_row_labels, column=3).font = SUMMARY_LABEL_FONT
    ws.cell(row=summary_row_labels, column=3).alignment = CENTER_WRAP
    ws.cell(row=summary_row_values, column=3).alignment = CENTER_WRAP
    
    ws.cell(row=summary_row_labels, column=4).value = "Not Relevant"
    ws.cell(row=summary_row_values, column=4).value = not_relevant_count
    ws.cell(row=summary_row_labels, column=4).fill = SUMMARY_NOT_RELEVANT_FILL
    ws.cell(row=summary_row_values, column=4).fill = SUMMARY_NOT_RELEVANT_FILL
    ws.cell(row=summary_row_labels, column=4).font = SUMMARY_LABEL_FONT
    ws.cell(row=summary_row_labels, column=4).alignment = CENTER_WRAP
    ws.cell(row=summary_row_values, column=4).alignment = CENTER_WRAP
    
    ws.cell(row=summary_row_labels, column=5).value = "Total Executed"
    ws.cell(row=summary_row_values, column=5).value = executed_count
    ws.cell(row=summary_row_labels, column=5).fill = SUMMARY_EXECUTED_FILL
    ws.cell(row=summary_row_values, column=5).fill = SUMMARY_EXECUTED_FILL
    ws.cell(row=summary_row_labels, column=5).font = SUMMARY_LABEL_FONT
    ws.cell(row=summary_row_labels, column=5).alignment = CENTER_WRAP
    ws.cell(row=summary_row_values, column=5).alignment = CENTER_WRAP
    
    chart_data_col_start = 16
    chart_labels_row = 1
    chart_data_row = 2
    
    ws.cell(row=chart_labels_row, column=chart_data_col_start).value = "Passed"
    ws.cell(row=chart_labels_row, column=chart_data_col_start + 1).value = "Failed"
    ws.cell(row=chart_labels_row, column=chart_data_col_start + 2).value = "Not Relevant"
    
    ws.cell(row=chart_data_row, column=chart_data_col_start).value = pass_count
    ws.cell(row=chart_data_row, column=chart_data_col_start + 1).value = fail_count
    ws.cell(row=chart_data_row, column=chart_data_col_start + 2).value = not_relevant_count
    
    pie = PieChart()
    pie.title = "Status Distribution"
    pie.width = 6.5
    pie.height = 11.0
    pie_labels = Reference(ws, min_col=chart_data_col_start, min_row=chart_labels_row, max_row=chart_labels_row, max_col=chart_data_col_start + 2)
    pie_data = Reference(ws, min_col=chart_data_col_start, min_row=chart_data_row, max_row=chart_data_row, max_col=chart_data_col_start + 2)
    pie.add_data(pie_data, titles_from_data=False)
    pie.set_categories(pie_labels)
    ws.add_chart(pie, "F1")
    
    bar = BarChart()
    bar.title = "Test Summary"
    bar.type = "col"
    bar.style = 10
    bar.width = 6.5
    bar.height = 11.0
    bar.y_axis.title = "Count"
    bar.x_axis.title = "Category"
    bar_labels = Reference(ws, min_col=chart_data_col_start, min_row=chart_labels_row, max_row=chart_labels_row, max_col=chart_data_col_start + 2)
    bar_data = Reference(ws, min_col=chart_data_col_start, min_row=chart_data_row, max_row=chart_data_row, max_col=chart_data_col_start + 2)
    bar.add_data(bar_data, titles_from_data=False)
    bar.set_categories(bar_labels)
    ws.add_chart(bar, "H1")
    
    header_row = 10
    
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.fill = HEADER_FILL  # Blue background
        cell.font = HEADER_FONT  # White, bold text
        cell.alignment = CENTER_WRAP
    
    # Enable Excel filters on header row
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(HEADERS))}{header_row}"
    
    # Freeze panes at first data row
    ws.freeze_panes = f"A{header_row + 1}"
    
    data_start_row = 11
    
    if not qs.exists():
        ws.cell(row=data_start_row, column=1).value = "NO DATA"
        ws.cell(row=data_start_row, column=1).alignment = CENTER_WRAP
    else:
        test_case_ids = list(qs.values_list('id', flat=True))
        executions = TestExecution.objects.filter(instance=active_instance, test_case_id__in=test_case_ids).select_related('test_case', 'user')
        
        exec_version_filters_list = [
            Q(sw_part_number=sw_num) & Q(app_sw_version=version)
            for sw_num, version in latest_versions.items()
        ]
        if exec_version_filters_list:
            exec_version_filter = reduce(or_, exec_version_filters_list)
            executions = executions.filter(exec_version_filter)
        
        latest_execution_map = {}
        for exec_obj in executions:
            sw_num = exec_obj.sw_part_number
            if sw_num in latest_versions and exec_obj.app_sw_version == latest_versions[sw_num]:
                key = (exec_obj.test_case.id, sw_num)
                if key not in latest_execution_map:
                    latest_execution_map[key] = exec_obj
        
        current_row = data_start_row
        
        for t in qs:
            # Get execution data from latest version only
            exec_status = ""
            exec_reports = ""
            exec_comments = ""
            
            sw_num = t.sw_part_number
            if sw_num and sw_num in latest_versions:
                exec_key = (t.id, sw_num)
                exec_obj = latest_execution_map.get(exec_key)
                if exec_obj:
                    exec_status = exec_obj.status or ""
                    exec_reports = exec_obj.reports or ""
                    exec_comments = exec_obj.comments or ""
            
            # Data Source Enforcement:
            # - TestCase → design/spec fields only
            # - TestExecution → status, reports, comments
            # - Missing execution → NOT EXECUTED
            if exec_status:
                final_status = exec_status
            else:
                # Missing execution → NOT EXECUTED (do not use TestCase.status)
                final_status = "NOT EXECUTED"
            
            # Use execution reports/comments if available, otherwise empty
            final_reports = exec_reports or ""
            final_comments = exec_comments or ""
            
            # Write data row
            # Always use base_test_case_id for export (version suffix never shown)
            display_test_case_id = t.base_test_case_id or t.test_case_id
            # Auto-derive requirement_id from base_test_case_id if requirement_id is empty
            display_requirement_id = get_requirement_id(t.base_test_case_id or t.test_case_id, t.requirement_id)
            row_data = [
                "",  # STRICT: sl_no does NOT exist - removed after hierarchy refactor
                t.sw_part_number,
                t.feature,
                display_requirement_id,
                t.requirement_description,
                display_test_case_id,  # Always base_test_case_id (no version suffix)
                t.test_case_summary,
                t.pre_conditions,
                t.inputs,
                t.periodic_time,
                t.test_steps,
                t.expected_result,
                final_status,
                final_reports,
                final_comments,
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.alignment = CENTER_WRAP
                
                # Feature column (column 3) - Light Yellow background
                if col_idx == 3:
                    cell.fill = FEATURE_FILL
                
                # Status column (column 13) - Color based on status
                if col_idx == 13:
                    status = (final_status or "").upper()
                    if status == "PASS":
                        cell.fill = STATUS_PASS_FILL  # Green
                    elif status == "FAIL":
                        cell.fill = STATUS_FAIL_FILL  # Red
                    elif status in ("NOT RELEVANT", "NOT_RELEVANT"):
                        cell.fill = STATUS_NOT_RELEVANT_FILL  # Yellow
                    else:
                        cell.fill = STATUS_NOT_EXECUTED_FILL  # Gray
            
            current_row += 1
    
    # Set column widths
    for col_idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = FIXED_WIDTH


# =====================================================
# MAIN BUILDER
# =====================================================

def build_testcase_export_workbook(
    sheet_filter="",
    sw="",
    app_sw_version="",
    feature="",
    query="",
    versions_list=None,
    latest_versions_only=False,
    version_selections=None,
    selected_features=None,
    selected_version_objs=None
):
    """
    Builds Excel workbook with structure:
    - Sheet 1: Dashboard (with Project Overview)
    - Sheet 2: History
    - Sheet 3+: Test Case Sheets (one per sheet_name)
    
    NEW FEATURE-BASED EXPORT:
    - selected_features: List of feature names to export (if provided, filters by these features)
    - selected_version_objs: List of TestCaseVersion objects to export (if provided, uses these)
    - If both provided, exports ONLY test cases matching selected features AND selected versions
    
    Parameters:
    - version_selections: Dict mapping SW Part Number to selected app_sw_version (backward compatibility)
    - selected_features: List of feature names to export (NEW - feature-based export)
    - selected_version_objs: List of TestCaseVersion objects to export (NEW - feature-based export)
    - sheet_filter: Filter by specific sheet (empty = all sheets)
    - sw: Filter by specific SW Part Number (empty = all SWs)
    - feature: Filter by feature (backward compatibility, overridden by selected_features)
    - query: Search query
    
    Only exports data from the active test instance.
    """
    from functools import reduce
    from operator import or_
    from django.db.models import Q
    
    # Get active instance - only export active instance data
    active_instance = get_active_instance()
    
    # Decide which sheets to export (ALWAYS export ALL sheets)
    sheet_names = list(
        SheetMeta.objects.values_list("sheet_name", flat=True)
    )
    if not sheet_names:
        sheet_names = list(
            TestCase.objects.filter(instance=active_instance).values_list("sheet_name", flat=True).distinct()
        )
    
    # Create workbook
    wb = openpyxl.Workbook()
    if wb.active:
        wb.remove(wb.active)
    
    # Determine which versions to use
    if selected_version_objs:
        # NEW: Use selected version objects (feature-based export)
        # Build version_selections dict from version objects
        selected_versions = {}
        for version_obj in selected_version_objs:
            sw_num = version_obj.sw_part_number
            version = version_obj.app_sw_version
            # Handle multiple versions per SW
            if sw_num in selected_versions:
                if isinstance(selected_versions[sw_num], str):
                    selected_versions[sw_num] = [selected_versions[sw_num], version]
                else:
                    selected_versions[sw_num].append(version)
            else:
                selected_versions[sw_num] = version
    elif version_selections:
        # Use selected versions per SW Part Number (backward compatibility)
        selected_versions = version_selections
    else:
        # Fallback: Get most recently created versions (backward compatibility)
        all_test_cases = TestCase.objects.filter(instance=active_instance, sheet_name__in=sheet_names)
        sw_part_numbers = list(all_test_cases.values_list('sw_part_number', flat=True).distinct())
        sw_part_numbers = [sw for sw in sw_part_numbers if sw]
        selected_versions = _get_most_recently_created_version(sw_part_numbers)
    
    # =================================================
    # SHEET 1: DASHBOARD (with Project Overview)
    # =================================================
    _write_dashboard_sheet(wb, sheet_names, selected_versions, active_instance)
    
    # =================================================
    # SHEET 2: HISTORY
    # =================================================
    _write_history_sheet(wb)
    
    # =================================================
    # SHEET 3+: TEST CASE SHEETS (ALL sheets)
    # =================================================
    for sheet_name in sheet_names:
        ws = wb.create_sheet(title=sheet_name)
        
        # Build queryset for this sheet (active instance only)
        qs = TestCase.objects.filter(instance=active_instance, sheet_name=sheet_name)
        
        # Filter to ONLY selected versions per SW Part Number
        version_filters_list = []
        for sw_num, version in selected_versions.items():
            if isinstance(version, list):
                # Multiple versions for same SW
                for v in version:
                    version_filters_list.append(Q(sw_part_number=sw_num) & Q(app_sw_version=v))
            else:
                version_filters_list.append(Q(sw_part_number=sw_num) & Q(app_sw_version=version))
        
        if version_filters_list:
            version_filter = reduce(or_, version_filters_list)
            qs = qs.filter(version_filter)
        
        # NEW: Filter by selected features if provided
        if selected_features:
            qs = qs.filter(feature__in=selected_features)
        elif feature:
            # Backward compatibility: use feature parameter
            qs = qs.filter(feature=feature)
        
        if query:
            search_filter = build_search_filters(query)
            qs = qs.filter(search_filter)
        
        # Order by id (primary key)
        qs = qs.order_by("id")
        
        # Write sheet data
        _write_test_case_sheet(ws, qs, selected_versions, active_instance)
    
    return wb
