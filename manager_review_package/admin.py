from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
from django.urls import reverse
import io
import openpyxl
from openpyxl.utils import get_column_letter
from .models import TestExecution, TestCase, SheetMeta, TestExecutionSnapshot
from testmanager.decorators import is_manager
from testmanager.services import get_active_instance
from testmanager.version_service import get_latest_versions


def _build_workbook_from_queryset(qs):
    """
    Returns an openpyxl Workbook with one sheet named 'TestCases' populated
    from queryset `qs`. This mirrors the export logic used in your views but
    keeps it small for admin exports.
    """
    headers = [
        "SL.NO", "SW Part Number", "Application SW Version", "Feature", "Requirement ID",
        "Requirement Description", "Test Case ID", "Test Case Summary",
        "Pre Conditions", "Inputs", "Periodic Time", "Test Steps",
        "Expected Result", "Status", "Reports", "Comments"
    ]

    wb = openpyxl.Workbook()
    # Get active sheet; ensure it's not None
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet(title="TestCases")
    else:
        ws.title = "TestCases"

    # pad first 7 blank rows to keep compatibility with your Excel format
    for _ in range(7):
        ws.append([""] * len(headers))
    ws.append(headers)

    for t in qs:
        ws.append([
            t.sl_no, t.sw_part_number, t.app_sw_version, t.feature, t.requirement_id,
            t.requirement_description, t.test_case_id, t.test_case_summary,
            t.pre_conditions, t.inputs, t.periodic_time, t.test_steps,
            t.expected_result, t.status, t.reports, t.comments
        ])

    # Auto-width columns: safer iteration using iter_cols()
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)), start=1):
        max_len = 0
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 5

    return wb


@admin.action(description="Export selected test cases to Excel")
def export_selected_as_excel(modeladmin, request, queryset):
    if not queryset.exists():
        modeladmin.message_user(request, "No test cases selected.", level="warning")
        return

    wb = _build_workbook_from_queryset(queryset)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=TestCases_Export.xlsx"
    return response


@admin.register(TestCase)
class TestCaseAdmin(admin.ModelAdmin):
    """
    Admin interface for Test Cases with version-wise grouping.
    
    STRICT VERSION LIFECYCLE CONTROL:
    - Managers: See all versions grouped by version number (latest first)
    - Non-managers: See only the latest active version (is_active=True)
    - Database structure enforces version isolation via unique_together constraints
    - Each version represents a complete, isolated execution dataset
    """
    list_display = (
        "sl_no_numeric",
        "test_case_id",
        "sheet_name",
        "sw_part_number",
        "app_sw_version",
        "feature",
        "status",
        "updated_at",
        "created_at",
    )
    list_display_links = ("test_case_id",)

    # -----------------------------
    # SORTING (CLICKABLE HEADERS)
    # -----------------------------
    # STRICT: Order by version (latest first) then by updated_at for version-wise grouping
    # Database structure enforces version isolation via unique_together constraints
    ordering = ("-app_sw_version", "-updated_at")
    
    class Media:
        css = {
            'all': ('testmanager/css/admin_custom.css',)
        }
    
    def get_queryset(self, request):
        """
        Role-based queryset filtering:
        - Managers: See all versions
        - Non-managers: See only latest active version (is_active=True)
        """
        qs = super().get_queryset(request)
        
        # Managers see all versions
        if is_manager(request.user) or request.user.is_superuser:
            return qs
        
        # Non-managers: Filter to only latest active versions
        active_instance = get_active_instance()
        if not active_instance:
            return qs.none()
        
        # Get latest active versions for all SW part numbers
        latest_versions = get_latest_versions(active_instance=active_instance)
        
        if not latest_versions:
            return qs.none()
        
        # Filter to only test cases with latest active versions
        from django.db.models import Q
        from functools import reduce
        from operator import or_
        
        version_filters = [
            Q(sw_part_number=sw_num) & Q(app_sw_version=version)
            for sw_num, version in latest_versions.items()
        ]
        
        if version_filters:
            qs = qs.filter(reduce(or_, version_filters))
        
        return qs.filter(instance=active_instance)

    def sl_no_numeric(self, obj):
        try:
            return int(obj.sl_no)
        except Exception:
            return obj.sl_no

    sl_no_numeric.short_description = "SL.NO"
    sl_no_numeric.admin_order_field = "sl_no"


    search_fields = (
        "test_case_id",
        "sw_part_number","app_sw_version",
        "feature",
        "requirement_id",
        "test_case_summary",
    )

    list_filter = (
        "app_sw_version",  # STRICT: Version filter FIRST for version-first structure
        "sheet_name",
        "sw_part_number",
        "status",
        "instance",  # Filter by instance for version isolation
        "updated_at",
        "created_at",
    )


    readonly_fields = ("created_at", "updated_at")
    list_per_page = 50
    date_hierarchy = "updated_at"

    actions = [export_selected_as_excel]

    fieldsets = [
        (None, {
            "fields": [
                "sheet_name",
                "sl_no",
                "sw_part_number",
                "app_sw_version",
                "feature",
                "requirement_id",
                "requirement_description",
                "test_case_id",
                "test_case_summary",
                "pre_conditions",
                "inputs",
                "periodic_time",
                "test_steps",
                "expected_result",
                "status",
                "reports",
                "comments",
            ]
        }),
        ("Timestamps", {"fields": ("created_at", "updated_at")}),
    ]

    def short_summary(self, obj):
        if obj.test_case_summary and len(obj.test_case_summary) > 40:
            return obj.test_case_summary[:37] + "..."
        return obj.test_case_summary
    short_summary.short_description = "Summary"

    def save_model(self, request, obj, form, change):
        if obj.sheet_name:
            obj.sheet_name = obj.sheet_name.upper()
        super().save_model(request, obj, form, change)


@admin.register(SheetMeta)
class SheetMetaAdmin(admin.ModelAdmin):
    list_display = ("sheet_name",)
    search_fields = ("sheet_name",)
    readonly_fields = ("headers",)

    
# ActivityLog removed - using TestExecution instead

@admin.register(TestExecution)
class TestExecutionAdmin(admin.ModelAdmin):
    """
    Admin interface for Test Executions with version-wise grouping.
    
    STRICT VERSION LIFECYCLE CONTROL:
    - Managers: See all execution data for all versions
    - Non-managers: See only execution data for latest active version (is_active=True)
    - Execution data is stored per version and never overwrites previous versions
    """
    list_display = (
        "test_case_info",
        "sw_part_number",
        "app_sw_version",
        "status_display",
        "user",
        "executed_at",
        "reports_preview",
        "comments_preview",
    )
    
    list_display_links = ("test_case_info",)
    
    ordering = ("-executed_at",)
    
    class Media:
        css = {
            'all': ('testmanager/css/admin_custom.css',)
        }
    
    def get_queryset(self, request):
        """
        Role-based queryset filtering:
        - Managers: See all execution data for all versions
        - Non-managers: See only execution data for latest active version (is_active=True)
        """
        qs = super().get_queryset(request)
        
        # Managers see all versions
        if is_manager(request.user) or request.user.is_superuser:
            return qs
        
        # Non-managers: Filter to only latest active versions
        active_instance = get_active_instance()
        if not active_instance:
            return qs.none()
        
        # Get latest active versions for all SW part numbers
        latest_versions = get_latest_versions(active_instance=active_instance)
        
        if not latest_versions:
            return qs.none()
        
        # Filter to only executions with latest active versions
        from django.db.models import Q
        from functools import reduce
        from operator import or_
        
        version_filters = [
            Q(sw_part_number=sw_num) & Q(app_sw_version=version)
            for sw_num, version in latest_versions.items()
        ]
        
        if version_filters:
            qs = qs.filter(reduce(or_, version_filters))
        
        return qs.filter(instance=active_instance)
    
    date_hierarchy = "executed_at"
    
    list_filter = (
        "status",
        "sw_part_number",
        "app_sw_version",
        "executed_at",
        "user",
    )

    search_fields = (
        "test_case__test_case_id",
        "test_case__test_case_summary",
        "test_case__sheet_name",
        "sw_part_number",
        "app_sw_version",
        "status",
        "reports",
        "comments",
        "user__username",
        "user__email",
    )
    
    readonly_fields = (
        "test_case",
        "user",
        "sw_part_number",
        "app_sw_version",
        "status",
        "reports",
        "comments",
        "executed_at",
    )
    
    list_per_page = 50
    
    fieldsets = [
        ("Test Case Information", {
            "fields": ["test_case"]
        }),
        ("Execution Details", {
            "fields": [
                "sw_part_number",
                "app_sw_version",
                "status",
                "reports",
                "comments",
            ]
        }),
        ("Execution Metadata", {
            "fields": [
                "user",
                "executed_at",
            ]
        }),
    ]
    
    def test_case_info(self, obj):
        """Display test case information"""
        if obj.test_case:
            return f"{obj.test_case.test_case_id} ({obj.test_case.sheet_name})"
        return "-"
    test_case_info.short_description = "Test Case"
    test_case_info.admin_order_field = "test_case__test_case_id"
    
    def status_display(self, obj):
        """Display status with color coding"""
        if not obj.status:
            return "-"
        status = obj.status.upper()
        if status == "PASS":
            return format_html('<span style="color: #28a745; font-weight: bold;">{}</span>', obj.status)
        elif status == "FAIL":
            return format_html('<span style="color: #dc3545; font-weight: bold;">{}</span>', obj.status)
        else:
            return format_html('<span style="color: #ffc107; font-weight: bold;">{}</span>', obj.status)
    status_display.short_description = "Status"
    status_display.admin_order_field = "status"
    
    def reports_preview(self, obj):
        """Show truncated reports"""
        if obj.reports:
            if len(obj.reports) > 50:
                return obj.reports[:50] + "..."
            return obj.reports
        return "-"
    reports_preview.short_description = "Reports"
    
    def comments_preview(self, obj):
        """Show truncated comments"""
        if obj.comments:
            if len(obj.comments) > 50:
                return obj.comments[:50] + "..."
            return obj.comments
        return "-"
    comments_preview.short_description = "Comments"
    
    def has_add_permission(self, request):
        """Prevent adding executions from admin - they should be created through the test execution flow"""
        return False
    
    def has_delete_permission(self, request, obj=None):
        """Only superusers can delete execution history"""
        return request.user.is_superuser


@admin.register(TestExecutionSnapshot)
class TestExecutionSnapshotAdmin(admin.ModelAdmin):
    """
    Admin interface for Test Execution Snapshots.
    
    Each export action creates a snapshot record with:
    - Version number
    - Export timestamp
    - Permanent export link (HTML)
    - Execution data (status, reports, comments)
    
    Snapshots are permanently stored and reusable - managers can share export links.
    """
    list_display = ('snapshot_name', 'sheet_name', 'sw_part_number', 'app_sw_version', 'export_link', 'total_test_cases', 'total_executed', 'total_passed', 'total_failed', 'exported_by', 'exported_at')
    
    class Media:
        css = {
            'all': ('testmanager/css/admin_custom.css',)
        }
    
    # STRICT: Sheet-wise grouping - order by sheet_name first, then version, then timestamp
    ordering = ('sheet_name', '-app_sw_version', '-exported_at')
    
    list_filter = ('sheet_name', 'app_sw_version', 'sw_part_number', 'exported_at', 'exported_by')
    search_fields = ('snapshot_name', 'sheet_name', 'sw_part_number', 'app_sw_version', 'notes', 'export_id')
    readonly_fields = ('exported_at', 'execution_data', 'total_test_cases', 'total_executed', 'total_passed', 'total_failed', 'total_not_executed', 'export_link_field')
    date_hierarchy = 'exported_at'
    
    fieldsets = (
        ('Snapshot Information', {
            'fields': ('snapshot_name', 'sheet_name', 'sw_part_number', 'app_sw_version', 'exported_by', 'exported_at')
        }),
        ('Export Link', {
            'fields': ('export_link_field',),
            'description': 'Permanent link to view this snapshot. Share this link to access the exported report.'
        }),
        ('Statistics', {
            'fields': ('total_test_cases', 'total_executed', 'total_passed', 'total_failed', 'total_not_executed')
        }),
        ('Execution Data', {
            'fields': ('execution_data',),
            'classes': ('collapse',)
        }),
        ('Notes', {
            'fields': ('notes',)
        }),
    )
    
    def export_link(self, obj):
        """Display clickable export link in list view"""
        if obj.export_id:
            url = reverse('export_html_snapshot', args=[obj.export_id])
            return format_html('<a href="{}" target="_blank">View Snapshot</a>', url)
        return "-"
    export_link.short_description = "Export Link"
    export_link.admin_order_field = "export_id"
    
    def export_link_field(self, obj):
        """Display export link in detail view"""
        if obj.export_id:
            url = reverse('export_html_snapshot', args=[obj.export_id])
            return format_html(
                '<p><strong>Permanent Export Link:</strong></p>'
                '<p><a href="{}" target="_blank" class="button">{}</a></p>'
                '<p><small>Share this link to provide access to the exported report.</small></p>',
                url, url
            )
        return "-"
    export_link_field.short_description = "Export Link"
    
    def get_queryset(self, request):
        """
        Role-based queryset filtering:
        - Managers: See all snapshots
        - Non-managers: See only snapshots for latest active version (is_active=True)
        """
        qs = super().get_queryset(request)
        
        # Managers see all snapshots
        if is_manager(request.user) or request.user.is_superuser:
            return qs
        
        # Non-managers: Filter to only latest active versions
        active_instance = get_active_instance()
        if not active_instance:
            return qs.none()
        
        # Get latest active versions for all SW part numbers
        latest_versions = get_latest_versions(active_instance=active_instance)
        
        if not latest_versions:
            return qs.none()
        
        # Filter to only snapshots with latest active versions
        from django.db.models import Q
        from functools import reduce
        from operator import or_
        
        version_filters = [
            Q(sw_part_number=sw_num) & Q(app_sw_version=version)
            for sw_num, version in latest_versions.items()
        ]
        
        if version_filters:
            qs = qs.filter(reduce(or_, version_filters))
        
        return qs.filter(instance=active_instance)
