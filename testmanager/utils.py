# testmanager/utils.py
import os
import logging
from typing import List, Any, Dict, Optional
from .models import ActivityLog
from .constants import (
    BULK_CREATE_BATCH_SIZE,
    PROJECT_OVERVIEW_START_ROW,
    PROJECT_OVERVIEW_LABEL_COL,
    PROJECT_OVERVIEW_VALUE_COL,
    PROJECT_OVERVIEW_MAX_EMPTY_ROWS,
)

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# ---------------------------------------------------------
# 1. BASIC NORMALIZATION & CLEANING
# ---------------------------------------------------------
def clean(value):
    """Convert any value into a clean string."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(value).strip()
    return str(value).strip()


def clean_slno(value):
    """
    Special cleaner for SL.NO:
    - 1.0 → "1"
    - None → ""
    - Strings are stripped
    """
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value).strip()
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def normalize(text):
    """
    Used to normalize Excel headers before mapping:
    "SW Part Number" -> "swpartnumber"
    """
    if not text:
        return ""
    return (
        str(text).lower()
        .replace(" ", "")
        .replace("-", "")
        .replace("(", "")
        .replace(")", "")
        .replace("_", "")
    )

# ---------------------------------------------------------
# 2. HEADER MAPPING
# ---------------------------------------------------------
HEADER_MAPPING = {
    "sl.no": "sl_no",
    "slno": "sl_no",

    "swpartnumber": "sw_part_number",
    "sw part number": "sw_part_number",

    "feature": "feature",

    "requirement id": "requirement_id",
    "requirementid(dng)": "requirement_id",
    "requirement description": "requirement_description",
    "requirementdescription": "requirement_description",

    "testcaseid": "test_case_id",
    "test case id": "test_case_id",

    "testcasesummary": "test_case_summary",
    "test case summary": "test_case_summary",

    "preconditions": "pre_conditions",
    "pre conditions": "pre_conditions",

    "inputs": "inputs",

    "periodictime": "periodic_time",
    "periodic time": "periodic_time",

    "teststeps": "test_steps",
    "test steps": "test_steps",

    "expected result": "expected_result",
    "expectedresult": "expected_result",

    "status": "status",
    "reports": "reports",
    "comments": "comments",
}


# ---------------------------------------------------------
# 3. HEADER ROW DETECTION
# ---------------------------------------------------------
def detect_header_row(ws):
    """
    Detects the header row in an Excel worksheet by trying row 7 first, then row 10.
    Uses partial matching: a row is valid if it contains at least 3 expected headers.
    
    Returns:
        tuple: (header_row, raw_headers, normalized_headers)
            - header_row: The detected row number (7 or 10)
            - raw_headers: List of raw header values (after clean())
            - normalized_headers: List of normalized header values (after normalize())
    
    Raises:
        ValueError: If neither row 7 nor row 10 contains valid headers.
    """
    header_map = {normalize(k): v for k, v in HEADER_MAPPING.items()}
    
    for candidate_row in [7, 10]:
        raw_headers = [clean(c.value) for c in ws[candidate_row]]
        normalized_headers = [normalize(h) for h in raw_headers]
        
        matched_count = 0
        for nh in normalized_headers:
            if nh in header_map:
                matched_count += 1
        
        if matched_count >= 3:
            return (candidate_row, raw_headers, normalized_headers)
    
    raise ValueError("Header row not found. Expected at row 7 or row 10.")


# ---------------------------------------------------------
# 4. SAFE WORKBOOK LOADER
# ---------------------------------------------------------
def safe_load_workbook(file_bytes, data_only=True):
    """
    Safe wrapper around openpyxl.load_workbook.
    Always logs error and raises the exception upward.
    """
    try:
        return load_workbook(file_bytes, data_only=data_only)
    except Exception as e:
        logger.exception("openpyxl failed to load workbook: %s", e)
        raise


# ---------------------------------------------------------
# 5. BULK CREATE WITH BATCHES
# ---------------------------------------------------------
def bulk_create_in_batches(model_class, instances: List[Any], batch_size=None):
    """
    Bulk create many ORM objects efficiently.
    
    RISK REMOVAL: Batch size now uses constant from constants.py instead of hardcoded value.
    
    Returns the number of rows actually added.
    """
    if batch_size is None:
        batch_size = BULK_CREATE_BATCH_SIZE
    
    total = 0
    for i in range(0, len(instances), batch_size):
        batch = instances[i:i + batch_size]
        model_class.objects.bulk_create(batch, batch_size=len(batch))
        total += len(batch)
    return total


# ---------------------------------------------------------
# 6. PROJECT OVERVIEW READER + DISK CACHE
# ---------------------------------------------------------
_PROJECT_OVERVIEW_CACHE: Dict[str, Optional[Any]] = {"mtime": None, "data": None}


def read_project_overview_from_excel(path: str) -> Dict[str, Any]:
    """
    Reads project overview from the first sheet of project_overview.xlsx.
    Excel layout assumptions:
        - Row 8+ contains key/value pairs:
            Column E = key
            Column F = value
    Returns a dict:
        {"project code": "XYZ", "checksum value": "...", ...}
    """
    result: Dict[str, Any] = {}

    if not os.path.exists(path):
        return result

    try:
        mtime = os.path.getmtime(path)
    except OSError:
        return result

    cache = _PROJECT_OVERVIEW_CACHE
    cached_data = cache.get("data")

    # Use cached version if file unchanged
    if cache.get("mtime") == mtime and isinstance(cached_data, dict):
        return cached_data

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        logger.exception("Failed reading project overview file: %s", e)
        return result

    # RISK REMOVAL: Using constants instead of hardcoded row numbers and column letters
    start_row = PROJECT_OVERVIEW_START_ROW
    label_col = PROJECT_OVERVIEW_LABEL_COL
    value_col = PROJECT_OVERVIEW_VALUE_COL
    max_empty = PROJECT_OVERVIEW_MAX_EMPTY_ROWS
    empty_count = 0

    for r in range(start_row, ws.max_row + 1):
        key = ws[f"{label_col}{r}"].value
        val = ws[f"{value_col}{r}"].value

        if key is None or (isinstance(key, str) and key.strip() == ""):
            empty_count += 1
            if empty_count >= max_empty:
                break
            continue

        empty_count = 0

        k = (
            str(key)
            .strip()
            .lower()
            .replace(":", "")
        )
        v = "" if val is None else str(val).strip()
        result[k] = v

    cache["mtime"] = mtime
    cache["data"] = result
    return result

def log_activity(user, action, reference="", remarks="", content_type="TestCase"):
    ActivityLog.objects.create(
        user=user,
        action=action,
        reference=reference,
        remarks=remarks,
        content_type=content_type
    )

def get_model_diff(old_obj, new_data, fields):
    diff = {}

    for field in fields:
        old_val = getattr(old_obj, field)
        new_val = new_data.get(field)

        if str(old_val) != str(new_val):
            diff[field] = {
                "old": old_val,
                "new": new_val
            }

    return diff


# ---------------------------------------------------------
# 7. REQUIREMENT ID DERIVATION
# ---------------------------------------------------------
def get_requirement_id(test_case_id, requirement_id=None):
    """
    Derives Requirement ID from Test Case ID if requirement_id is empty.
    
    Test Case ID format: T_<REQUIREMENT_ID>_<SEQUENCE>
    Example: T_123124_123 → Requirement ID = 123124
    
    Args:
        test_case_id: The test case ID string (e.g., "T_123124_123")
        requirement_id: The stored requirement_id (may be None or empty)
    
    Returns:
        The requirement_id if it exists and is not empty,
        otherwise the derived requirement_id from test_case_id,
        or empty string if format doesn't match.
    """
    # If requirement_id is provided and not empty, use it
    if requirement_id and str(requirement_id).strip():
        return str(requirement_id).strip()
    
    # If test_case_id is empty or None, return empty string
    if not test_case_id or not str(test_case_id).strip():
        return ""
    
    test_case_id_str = str(test_case_id).strip()
    
    # Split by underscore
    parts = test_case_id_str.split("_")
    
    # Need at least 3 parts: T, REQUIREMENT_ID, SEQUENCE
    if len(parts) < 3:
        return ""
    
    # Extract the middle value (index 1) as Requirement ID
    derived_requirement_id = parts[1].strip()
    
    # Return empty string if derived value is empty
    return derived_requirement_id if derived_requirement_id else ""