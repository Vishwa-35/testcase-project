"""
Excel Import Views

This module contains views for uploading and importing Excel files with test cases.
"""

import base64
import io
import time
import logging
from django.db import transaction

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import permission_required, login_required
from django.utils import timezone

from ..models import TestCase, SheetMeta, SWVersionMapping, ActivityLog, TestCaseVersion, TestCaseSheet
from ..services import get_active_instance
from ..utils import (
    clean, clean_slno, normalize, HEADER_MAPPING,
    safe_load_workbook, bulk_create_in_batches, detect_header_row,
)
from ..constants import (
    EXCEL_HEADER_ROW,
    EXCEL_DATA_START_ROW,
    EXCEL_SHEETS_TO_SKIP,
    NOID_PREFIX,
    ACTIVITY_ACTION_IMPORT,
)

logger = logging.getLogger(__name__)


# -------------------------------------------------------------------------------------
# UPLOAD → SELECT SHEET
# -------------------------------------------------------------------------------------
@login_required
def upload_excel(request):
    """
    Excel Upload View
    
    PERMISSIONS: Manager only (superuser)
    TESTERS CANNOT upload Excel.
    """
    # PERMISSION CHECK: Manager only
    if not request.user.is_superuser:
        messages.error(request, "Permission denied. Only managers can upload Excel files.")
        return redirect("home")
    
    if request.method == "POST":
        file = request.FILES.get("excel_file")
        if not file or not file.name.endswith(".xlsx"):
            messages.error(request, "Upload a valid .xlsx file.")
            return redirect("upload_excel")

        file_data = base64.b64encode(file.read()).decode()

        try:
            wb = safe_load_workbook(io.BytesIO(base64.b64decode(file_data)))
        except Exception:
            messages.error(request, "Uploaded file could not be read.")
            return redirect("upload_excel")

        # RISK REMOVAL: Using constant for sheets to skip
        sheet_names = wb.sheetnames[EXCEL_SHEETS_TO_SKIP:]
        return render(request, "testmanager/select_sheet.html", {
            "file_data": file_data,
            "sheet_names": sheet_names,
            "next": request.GET.get("next", "")
        })

    return render(request, "testmanager/upload.html")


@login_required
def input_versions(request):
    """
    Extract SW part numbers from selected sheets and ask for versions
    
    PERMISSIONS: Manager only (superuser)
    TESTERS CANNOT upload Excel.
    """
    # PERMISSION CHECK: Manager only
    if not request.user.is_superuser:
        messages.error(request, "Permission denied. Only managers can upload Excel files.")
        return redirect("home")
    
    if request.method != "POST":
        return redirect("upload_excel")

    file_data = request.POST.get("file_data")
    selected_sheets = request.POST.getlist("sheet_names")

    if not file_data or not selected_sheets:
        messages.error(request, "No sheets selected.")
        return redirect("upload_excel")

    file_bytes = base64.b64decode(file_data)

    try:
        wb = safe_load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        logger.exception("Workbook load failed: %s", exc)
        messages.error(request, "Workbook could not be opened.")
        return redirect("upload_excel")

    # Extract unique SW part numbers from selected sheets
    sw_part_numbers = set()
    for sheet_name in selected_sheets:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        try:
            HEADER_ROW, raw_headers, normalized_headers = detect_header_row(ws)
            DATA_START = HEADER_ROW + 1
        except ValueError as e:
            logger.exception("Header detection failed for sheet %s: %s", sheet_name, e)
            continue

        header_map = {normalize(k): v for k, v in HEADER_MAPPING.items()}

        # Find SW Part Number column index
        sw_index = None
        for idx, nh in enumerate(normalized_headers):
            if nh in header_map and header_map[nh] == "sw_part_number":
                sw_index = idx
                break

        if sw_index is not None:
            # Extract SW part numbers from data rows
            for row in ws.iter_rows(min_row=DATA_START, values_only=True):
                if row and len(row) > sw_index:
                    sw_val = clean(row[sw_index])
                    if sw_val:
                        sw_part_numbers.add(sw_val)

    # PART 2: Get existing versions for these SW part numbers from active instance only
    active_instance = get_active_instance()
    existing_mappings = {}
    for mapping in SWVersionMapping.objects.filter(
        instance=active_instance,
        sw_part_number__in=sw_part_numbers
    ):
        existing_mappings[mapping.sw_part_number] = mapping.version

    return render(request, "testmanager/input_versions.html", {
        "file_data": file_data,
        "selected_sheets": selected_sheets,
        "sw_part_numbers": sorted(sw_part_numbers),
        "existing_mappings": existing_mappings,
        "next": request.POST.get("next", "")
    })

# -------------------------------------------------------------------------------------
# IMPORT SELECTED SHEETS (MULTIPLE)
# -------------------------------------------------------------------------------------
@login_required
def import_excel(request):
    """
    Import Excel file with test cases
    
    PERMISSIONS: Manager only (superuser)
    TESTERS CANNOT upload Excel.
    
    VERSION RULES:
    - Version asked ONCE per sw_part_number in popup
    - Apply same version to all rows of that SW Part Number
    - Save rows ONLY to active instance + entered version
    """
    # PERMISSION CHECK: Manager only
    if not request.user.is_superuser:
        messages.error(request, "Permission denied. Only managers can upload Excel files.")
        return redirect("home")
    
    if request.method != "POST":
        return redirect("upload_excel")

    file_data = request.POST.get("file_data")
    selected_sheets = request.POST.getlist("sheet_names")
    
    # PART 2 & 3: Get version mappings from popup (SINGLE SOURCE OF TRUTH)
    # Popup submit must send: {sw_part_number: version_value}
    version_mappings = {}
    for key, value in request.POST.items():
        if key.startswith("version_"):
            sw_part_number = key.replace("version_", "")
            version = value.strip()
            if version:
                version_mappings[sw_part_number] = version

    # PART 3: Store version mappings with instance
    # Create or update SWVersionMapping with instance = active_instance
    # Set is_active=True for new versions, is_active=False for previous versions
    active_instance = get_active_instance()
    for sw_part_number, version in version_mappings.items():
        # Set all previous versions for this SW part number to is_active=False
        SWVersionMapping.objects.filter(
            instance=active_instance,
            sw_part_number=sw_part_number
        ).update(is_active=False)
        
        # Create or update with is_active=True
        SWVersionMapping.objects.update_or_create(
            instance=active_instance,  # PART 3: Bind to active instance
            sw_part_number=sw_part_number,
            defaults={"version": version, "is_active": True, "updated_at": timezone.now()}
        )

    if not file_data or not selected_sheets:
        messages.error(request, "No sheets selected.")
        return redirect("upload_excel")

    file_bytes = base64.b64decode(file_data)

    try:
        wb = safe_load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        logger.exception("Workbook load failed: %s", exc)
        messages.error(request, "Workbook could not be opened.")
        return redirect("upload_excel")

    # PART 4: Get active instance - all imported test cases must belong to active instance
    active_instance = get_active_instance()
    
    # PART 4: Fetch version from SWVersionMapping for each sw_part_number
    # Excel import must FAIL if version mapping missing for a SW part number
    version_mapping_dict = {}
    for sw_part_number, version in version_mappings.items():
        version_mapping_dict[sw_part_number] = version
    
    added_total = 0
    # Check existing test cases using (base_test_case_id + app_sw_version + instance) for duplicate detection
    existing_test_cases = TestCase.objects.filter(instance=active_instance).values_list(
        "base_test_case_id", "app_sw_version", flat=False
    )
    existing_keys = {(base_id, version) for base_id, version in existing_test_cases if base_id}
    noid_counter = int(time.time() * 1000)
    all_instances = []
    missing_versions = set()  # Track SW part numbers missing version mappings

    # CRITICAL: Each Excel worksheet MUST create one TestCaseSheet per worksheet
    # Rows from one worksheet MUST NOT mix with other sheets
    for sheet_name in selected_sheets:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        clean_sheet = sheet_name.upper()

        try:
            HEADER_ROW, raw_headers, normalized_headers = detect_header_row(ws)
            DATA_START = HEADER_ROW + 1
        except ValueError as e:
            logger.exception("Header detection failed for sheet %s: %s", sheet_name, e)
            messages.error(request, f"Header detection failed for sheet '{sheet_name}': {e}")
            continue

        header_map = {normalize(k): v for k, v in HEADER_MAPPING.items()}

        mapped_fields = []
        mapped_indexes = []
        for idx, nh in enumerate(normalized_headers):
            if nh in header_map:
                mapped_fields.append(header_map[nh])
                mapped_indexes.append(idx)

        added_sheet = 0
        
        # CRITICAL: Get unique SW part numbers from this sheet to create TestCaseVersion and TestCaseSheet
        sheet_sw_part_numbers = set()
        for row in ws.iter_rows(min_row=DATA_START, values_only=True):
            if not any(row):
                continue
            row = list(row)
            row_data = {}
            for idx, field in zip(mapped_indexes, mapped_fields):
                val = row[idx] if idx < len(row) else ""
                row_data[field] = clean_slno(val) if field == "sl_no" else clean(val)
            sw_part_number = row_data.get("sw_part_number", "").strip()
            if sw_part_number and sw_part_number in version_mapping_dict:
                sheet_sw_part_numbers.add(sw_part_number)
        
        # CRITICAL: Create TestCaseVersion and TestCaseSheet for each SW part number in this sheet
        # Each worksheet creates one TestCaseSheet per SW part number (version-bound)
        sheet_objects = {}  # Track TestCaseSheet objects by (sw_part_number, version)
        for sw_part_number in sheet_sw_part_numbers:
            mapped_version = version_mapping_dict.get(sw_part_number, "")
            if not mapped_version:
                continue
            
            # Get or create TestCaseVersion
            version_obj, _ = TestCaseVersion.objects.get_or_create(
                instance=active_instance,
                sw_part_number=sw_part_number,
                app_sw_version=mapped_version,
                defaults={"is_active": True, "is_locked": False}
            )
            
            # CRITICAL: Create one TestCaseSheet per worksheet for this version
            sheet_obj, _ = TestCaseSheet.objects.get_or_create(
                version=version_obj,
                sheet_name=clean_sheet
            )
            sheet_objects[(sw_part_number, mapped_version)] = sheet_obj

        # Process rows for this sheet
        # Track sl_no per sw_part_number for recalculation
        sw_sl_no_counters = {}  # Map sw_part_number -> current max sl_no
        
        for row in ws.iter_rows(min_row=DATA_START, values_only=True):
            if not any(row):
                continue

            row = list(row)
            row_data = {}

            for idx, field in zip(mapped_indexes, mapped_fields):
                val = row[idx] if idx < len(row) else ""
                # DO NOT use sl_no from Excel - we'll recalculate it per sw_part_number
                row_data[field] = clean(val) if field != "sl_no" else ""

            # Read base_test_case_id from Excel (this is the original ID without version suffix)
            base_tcid = (row_data.get("test_case_id") or "").strip()
            if base_tcid.startswith("="):
                base_tcid = ""

            # RISK REMOVAL: Using constant for NOID prefix
            if not base_tcid:
                noid_counter += 1
                base_tcid = f"{NOID_PREFIX}{noid_counter}"

            # PART 4: Get sw_part_number from row_data
            sw_part_number = row_data.get("sw_part_number", "").strip()
            
            # PART 4: Excel import must FAIL if version mapping missing for a SW part number
            if sw_part_number and sw_part_number not in version_mapping_dict:
                missing_versions.add(sw_part_number)
                continue  # Skip this row - version mapping missing
            
            # PART 4: Explicitly set version from SWVersionMapping
            # IGNORE: version inside Excel file, old TestCase versions, request version params
            mapped_version = version_mapping_dict.get(sw_part_number, "")
            
            # Check for duplicate using (base_test_case_id + app_sw_version + instance)
            if (base_tcid, mapped_version) in existing_keys:
                continue  # Skip duplicate
            
            # Generate versioned test_case_id
            versioned_tcid = f"{base_tcid}_{mapped_version}" if mapped_version else base_tcid
            
            # CRITICAL: Recalculate sl_no per sw_part_number (ignore Excel sl_no)
            # sl_no is scoped ONLY to sw_part_number (not sheet/version/feature)
            from ..utils import get_next_sl_no_for_sw_part_number
            # Get next sl_no for this sw_part_number, considering already processed rows in this batch
            if sw_part_number not in sw_sl_no_counters:
                # First time seeing this sw_part_number in this batch - get max from DB
                from django.db.models import Max
                max_result = TestCase.objects.filter(
                    instance=active_instance,
                    sw_part_number=sw_part_number
                ).exclude(sl_no__isnull=True).exclude(sl_no__exact="").aggregate(
                    max_sl_no=Max('sl_no')
                )
                max_sl_no = max_result.get('max_sl_no')
                try:
                    sw_sl_no_counters[sw_part_number] = int(max_sl_no) if max_sl_no else 0
                except (ValueError, TypeError):
                    sw_sl_no_counters[sw_part_number] = 0
            
            # Increment counter for this sw_part_number
            sw_sl_no_counters[sw_part_number] += 1
            row_data["sl_no"] = str(sw_sl_no_counters[sw_part_number])
            
            # CRITICAL: sl_no must be stored in TestCase (master definition)
            # sl_no must be immutable across versions
            row_data["sheet_name"] = clean_sheet
            row_data["base_test_case_id"] = base_tcid  # Store original ID
            row_data["test_case_id"] = versioned_tcid  # Store versioned ID
            existing_keys.add((base_tcid, mapped_version))  # Track for duplicate detection
            
            # Remove app_sw_version from row_data if present (we use mapped version instead)
            row_data.pop("app_sw_version", None)
            
            # PART 4: Assign to active instance and set version from mapping
            row_data["instance"] = active_instance
            row_data["app_sw_version"] = mapped_version  # PART 4: Use version from SWVersionMapping
            # PART 4: Execution data starts empty
            row_data["status"] = ""
            row_data["reports"] = ""
            row_data["comments"] = ""

            all_instances.append(TestCase(**row_data))
            added_sheet += 1

        # CRITICAL: Update SheetMeta for backward compatibility
        SheetMeta.objects.update_or_create(sheet_name=clean_sheet, defaults={"headers": raw_headers})
        added_total += added_sheet

    # PART 4: Excel import must FAIL if version mapping missing for a SW part number
    if missing_versions:
        messages.error(
            request,
            f"Import failed: Version mapping missing for SW Part Number(s): {', '.join(sorted(missing_versions))}. "
            f"Please enter versions for all SW Part Numbers in the popup."
        )
        return redirect("upload_excel")
    
    if all_instances:
        created_count = bulk_create_in_batches(TestCase, all_instances)
        added_total = created_count

    if added_total == 0:
        messages.warning(request, "No new rows imported (all duplicates).")
    else:
        messages.success(request, f"Import completed — {added_total} rows added.")

    # RISK REMOVAL: Using constant for action string
    ActivityLog.objects.create(
        user=request.user,
        action=ACTIVITY_ACTION_IMPORT,
        reference=sheet_name,
        remarks=f"{added_total} test cases imported",
        content_type="SheetMeta"
    )

    next_url = request.POST.get("next") or request.GET.get("next")
    if next_url:
        return redirect(next_url)
    return redirect("testcase_list")


# -------------------------------------------------------------------------------------
# FULL IMPORT (ALL SHEETS FROM 3RD ONWARD)
# -------------------------------------------------------------------------------------
@permission_required("testmanager.change_testcase", raise_exception=True)
def import_full_excel(request):
    if request.method != "POST":
        return redirect("upload_excel")

    file = request.FILES.get("excel_file")
    if not file or not file.name.endswith(".xlsx"):
        messages.error(request, "Upload a valid .xlsx file.")
        return redirect("upload_excel")

    try:
        from openpyxl import load_workbook
        wb = load_workbook(file, data_only=True)
    except Exception as exc:
        logger.exception("Full import failed: %s", exc)
        messages.error(request, "Excel cannot be opened.")
        return redirect("upload_excel")

    # RULE 1: Get active instance - all imported test cases must belong to active instance
    active_instance = get_active_instance()
    
    existing_ids = set(TestCase.objects.filter(instance=active_instance).values_list("test_case_id", flat=True))
    noid_counter = int(time.time() * 1000)
    all_instances = []
    total_added = 0

    # RISK REMOVAL: Using constant for sheets to skip
    for sheet_name in wb.sheetnames[EXCEL_SHEETS_TO_SKIP:]:
        ws = wb[sheet_name]
        clean_sheet = sheet_name.upper()

        try:
            HEADER_ROW, raw_headers, normalized_headers = detect_header_row(ws)
            DATA_START = HEADER_ROW + 1
        except ValueError as e:
            logger.exception("Header detection failed for sheet %s: %s", sheet_name, e)
            messages.error(request, f"Header detection failed for sheet '{sheet_name}': {e}")
            continue

        header_map = {normalize(k): v for k, v in HEADER_MAPPING.items()}

        mapped_fields = []
        mapped_indexes = []

        for idx, nh in enumerate(normalized_headers):
            if nh in header_map:
                mapped_fields.append(header_map[nh])
                mapped_indexes.append(idx)

        cnt = 0

        for row in ws.iter_rows(min_row=DATA_START, values_only=True):
            if not any(row):
                continue

            row = list(row)
            row_data = {}

            for idx, field in zip(mapped_indexes, mapped_fields):
                val = row[idx] if idx < len(row) else ""
                row_data[field] = clean_slno(val) if field == "sl_no" else clean(val)

            tcid = (row_data.get("test_case_id") or "").strip()
            if tcid.startswith("="):
                tcid = ""

            if tcid and tcid in existing_ids:
                continue

            # RISK REMOVAL: Using constant for NOID prefix
            if not tcid:
                noid_counter += 1
                tcid = f"{NOID_PREFIX}{noid_counter}"

            row_data["sheet_name"] = clean_sheet
            row_data["test_case_id"] = tcid
            existing_ids.add(tcid)
            
            # RULE 1: Assign to active instance
            row_data["instance"] = active_instance
            # RULE 4: Execution data starts empty
            row_data["status"] = ""
            row_data["reports"] = ""
            row_data["comments"] = ""

            all_instances.append(TestCase(**row_data))
            cnt += 1

        SheetMeta.objects.update_or_create(sheet_name=clean_sheet, defaults={"headers": raw_headers})
        total_added += cnt

    if all_instances:
        total_added = bulk_create_in_batches(TestCase, all_instances)

    if total_added == 0:
        messages.warning(request, "No new rows imported.")
    else:
        messages.success(request, f"Full import completed — {total_added} rows added.")

    next_url = request.POST.get("next") or request.GET.get("next")
    if next_url:
        return redirect(next_url)
    return redirect("testcase_list")

